import openpyxl
import datetime


if __name__ == '__main__':
    filename = "Data"
    path = filename + ".xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    cell_obj = sheet_obj.cell(row=1, column=1)
    n = sheet_obj.max_row
    print(n)
    sheet_title = sheet_obj.title
    for i in range(2, n + 1):
        x = sheet_obj.cell(row=i, column=1).value
        print(x, end=' ---> ')

        x = x[1:-1]

        #LOOK AT THERE, THIS IS THE MAGIC
        x = datetime.datetime.strptime(x, "%d-%b-%Y %H:%M:%S").time()

        print(x)
        sheet_obj.cell(row=i, column=1).value = x

    for i in range(2, n + 1):
        x = sheet_obj.cell(row=i, column=2).value
        print(x, end=' ---> ')
        x = float(x)
        print(x)
        sheet_obj.cell(row=i, column=2).value = x

    wb_obj.save(filename + "_DONE.xlsx")



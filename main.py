import functools as functools
import sys
import time
from datetime import datetime

import numpy as np
import openpyxl
from MainWindow import Ui_MainWindow
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QFileDialog, QMessageBox

import sunpy.timeseries as ts
from sunpy.net import Fido, attrs as a
import pyqtgraph as pg


functools.lru_cache(None)

print("_" * 100)


class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self, *args, obj=None, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ui.inputbutton.clicked.connect(self.choose_txt)
        self.ui.compare_button.clicked.connect(self.compare_click)
        self.ui.input_button_sun.clicked.connect(self.update_sun_activity)

        self.start_data = None
        self.end_data = None
        self.x_values = []

        plot_db = self.ui.grap_db
        plot_db.setLabel(axis='left', text='Мощность')
        plot_db.setLabel(axis='bottom', text='Время')

        plot_sun = self.ui.graph_sun
        plot_sun.setLabel(axis='left', text='Мощность солнца')
        plot_sun.setLabel(axis='bottom', text='Время')

        plot_out = self.ui.graph_out
        plot_out.setLabel(axis='left', text='Коэффициент корреляции')
        plot_out.setLabel(axis='bottom', text='Время')

    filepath_txt = ["", ""]
    filepath_sun_txt = ["", ""]
    successful_updated = False

    data_list_db_X = []
    data_list_db_Y = []

    corr_list_db_Y = []

    data_list_sun_X = []
    data_list_sun_Y = []

    corr_list_sun_Y = []

    range_1 = 0
    range_2 = 0
    range_end = 0

    _filepath_txt = None
    excel1 = None
    lists1 = None
    lists_num1 = None

    lists_combo = None

    _filepath_sun_txt = None
    excel2 = None
    lists2 = None
    lists_num2 = None

    def ClearDB(self):
        self.data_list_db_X.clear()
        self.data_list_db_Y.clear()
        self.ui.grap_db.clear()

    def ClearSun(self):
        self.data_list_sun_X.clear()
        self.data_list_sun_Y.clear()
        self.ui.graph_sun.clear()

    def ClearCompare(self):
        self.corr_list_db_Y.clear()
        self.corr_list_sun_Y.clear()
        self.ui.graph_out.clear()

    def read_selected_sheet(self, index):
        self.ClearDB()
        self.set_db_graph(0, 0, True)

        if not self.lists_combo:
            return
        else:
            selected_sheet_name = self.lists_combo[index]  # Получаем имя выбранного листа по индексу
            selected_sheet = self.excel1[selected_sheet_name]  # Получаем выбранный лист

        self.data_list_db_X.clear()
        self.data_list_db_Y.clear()
        previous_value = None  # Инициализируем переменную для хранения значения предыдущей ячейки столбца A
        previous_B_values = []

        for row in selected_sheet.iter_rows(min_row=2, max_row=selected_sheet.max_row, min_col=1, max_col=2):
            if all(cell.value is not None for cell in row):
                try:
                    current_value = row[0].value  # Получаем значение текущей ячейки столбца A
                    current_B_value = row[1].value  # Получаем значение текущей ячейки столбца B

                    # Преобразуем значение B в числовой формат, если оно не является числом
                    if not isinstance(current_B_value, (int, float)):
                        current_B_value = float(current_B_value)

                    if current_value == previous_value:  # Проверяем, совпадают ли текущее и предыдущее значения столбца A
                        previous_B_values.append(current_B_value)  # Добавляем значение B в список
                    else:
                        if previous_value is not None:  # Проверяем, было ли уже какое-то значение столбца A
                            # Если было, усредняем значения B и добавляем в список
                            avg_B = sum(previous_B_values) / len(previous_B_values)
                            self.data_list_db_Y.append(avg_B)
                            self.data_list_db_X.append(self.to_UNIX(previous_value))
                        # Обновляем значения для следующей итерации
                        previous_value = current_value
                        previous_B_values = [current_B_value]
                except Exception as e:
                    self.data_list_db_X.clear()  # Очищаем список данных X в случае ошибки
                    self.data_list_db_Y.clear()  # Очищаем список данных Y в случае ошибки

                    # Создание и отображение сообщения об ошибке
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Critical)
                    msg.setText("Ошибка")
                    msg.setInformativeText(
                        f'Не удалось конвертировать время в UNIX формат. Строка: {row}')
                    msg.setWindowTitle("Конвертация данных")
                    msg.exec_()
                    break  # Прерываем выполнение цикла в случае ошибки
            else:
                self.show_warning_message("Неправильные строки")
                break  # Прерываем выполнение цикла, если есть пустые ячейки в строке

        # Добавляем последнее значение
        if previous_value is not None:
            avg_B = sum(previous_B_values) / len(previous_B_values)
            self.data_list_db_Y.append(avg_B)
            self.data_list_db_X.append(self.to_UNIX(previous_value))

        if len(self.data_list_db_X) != 0 and len(self.data_list_db_Y) != 0:
            self.data_list_db_X.pop()
            self.data_list_db_Y.pop()

        self.start_data = selected_sheet.cell(row=2, column=1).value
        self.end_data = selected_sheet.cell(row=selected_sheet.max_row, column=1).value

        print("CubeSat")
        print(f"Выбран лист: {selected_sheet_name}")
        print("Количество строк:", len(self.data_list_db_Y))
        print("data_list_db_X:", self.data_list_db_X)
        print("data_list_db_Y:", self.data_list_db_Y)
        print("Первая строка:", self.start_data)
        print("Последняя строка:", self.end_data)
        print("_" * 100)

        try:
            self.data_list_db_X = list(map(float, self.data_list_db_X))
            self.data_list_db_Y = list(map(float, self.data_list_db_Y))
        except ValueError:
            self.show_critical_message("Неправильные строки", "Вероятно часть строк не соответствует нужному формату!")

        self.ui.file_path_window.setPlainText("Выбран файл мощности по пути: " + str(self._filepath_txt))
        self.set_db_graph(self.data_list_db_X, self.data_list_db_Y, False)

    def to_UNIX(self, date_string):
        date_string = date_string.strip("'")
        date_components = date_string.split('-')

        months = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10,
                  'Nov': 11, 'Dec': 12}
        month = months[date_components[1]]

        year = int(date_components[2][:4])
        time_components = date_components[2].split()
        time_parts = time_components[1].split(':')
        hour = int(time_parts[0])
        minute = int(time_parts[1])
        second = int(time_parts[2])

        datetime_obj = datetime(year, month, int(date_components[0]), hour, minute, second)
        unix_time = datetime_obj.timestamp()

        return unix_time

    def to_sunPy_time(self, date_string):
        date_string = date_string.strip("'")
        date_components = date_string.split('-')

        months = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10,
                  'Nov': 11, 'Dec': 12}
        month = months[date_components[1]]

        year = int(date_components[2][:4])
        time_components = date_components[2].split()
        time_parts = time_components[1].split(':')
        hour = int(time_parts[0])
        minute = int(time_parts[1])
        second = int(time_parts[2])

        datetime_obj = datetime(year, month, int(date_components[0]), hour, minute, second)
        unix_time = datetime_obj.timestamp()

        return datetime.utcfromtimestamp(unix_time).strftime('%Y-%m-%dT%H:%M:%S')

    def choose_txt(self):
        self.ClearDB()
        self._filepath_txt = QFileDialog.getOpenFileName(self, str("Загрузить .xlsx мощности"), "/",
                                                         str("xlsx (*.xlsx)"))

        if not self._filepath_txt[0]:
            self.excel1 = None
            self.lists1 = None
            self.filepath_txt = self._filepath_txt
            self.lists_combo = []
            self.ui.comboBox.clear()
            print("Выбран пустой файл или ничего не выбрано!")
            print("_" * 100)
            return

        self.excel1 = openpyxl.open(self._filepath_txt[0], read_only=True)
        self.lists1 = self.excel1.worksheets
        self.lists_num1 = 0

        self.lists_combo = self.excel1.sheetnames
        self.ui.comboBox.clear()
        self.ui.comboBox.addItems(self.lists_combo)
        self.ui.comboBox.currentIndexChanged.connect(self.read_selected_sheet)
        self._filepath_txt = self._filepath_txt[0]

        self.data_list_db_X = []
        self.data_list_db_Y = []

        previous_value = None
        previous_B_values = []

        for row in self.lists1[self.lists_num1].iter_rows(min_row=2, max_row=self.lists1[self.lists_num1].max_row, min_col=1, max_col=2):
            if all(cell.value is not None for cell in row):
                try:
                    current_value = row[0].value  # Получаем значение текущей ячейки столбца A
                    current_B_value = row[1].value  # Получаем значение текущей ячейки столбца B

                    # Преобразуем значение B в числовой формат, если оно не является числом
                    if not isinstance(current_B_value, (int, float)):
                        current_B_value = float(current_B_value)

                    if current_value == previous_value:  # Проверяем, совпадают ли текущее и предыдущее значения столбца A
                        previous_B_values.append(current_B_value)  # Добавляем значение B в список
                    else:
                        if previous_value is not None:  # Проверяем, было ли уже какое-то значение столбца A
                            # Если было, усредняем значения B и добавляем в список
                            avg_B = sum(previous_B_values) / len(previous_B_values)
                            self.data_list_db_Y.append(avg_B)
                            self.data_list_db_X.append(self.to_UNIX(previous_value))
                        # Обновляем значения для следующей итерации
                        previous_value = current_value
                        previous_B_values = [current_B_value]
                except:
                    self.data_list_db_X = []
                    self.data_list_db_Y = []

                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Critical)
                    msg.setText("Ошибка")
                    msg.setInformativeText(
                        f'Не удалось конвертировать время в UNIX формат. Строка: {row}')
                    msg.setWindowTitle("Конвертация данных")
                    msg.exec_()
                    break
            else:
                self.show_warning_message("Неправильные строки")
                break

        # Добавляем последнее значение
        if previous_value is not None:
            avg_B = sum(previous_B_values) / len(previous_B_values)
            self.data_list_db_Y.append(avg_B)
            self.data_list_db_X.append(self.to_UNIX(previous_value))

        if len(self.data_list_db_X) != 0 and len(self.data_list_db_Y) != 0:
            self.data_list_db_X.pop()
            self.data_list_db_Y.pop()

        self.start_data = self.lists1[self.lists_num1].cell(row=2, column=1).value
        self.end_data = self.lists1[self.lists_num1].cell(row=self.lists1[self.lists_num1].max_row, column=1).value

        print("CubeSat")
        print(f"Выбран файл: {self._filepath_txt}")
        print("Количество строк:", len(self.data_list_db_Y))
        print("data_list_db_X", self.data_list_db_X)
        print("data_list_db_Y", self.data_list_db_Y)
        print("Первая строка:", self.start_data)
        print("Последняя строка:", self.end_data)
        print("_" * 100)

        try:
            self.data_list_db_X = list(map(float, self.data_list_db_X))
            self.data_list_db_Y = list(map(float, self.data_list_db_Y))
        except ValueError:
            self.show_critical_message("Неправильные строки", "Вероятно часть строк не соответствует нужному формату!")

        if self._filepath_txt[0]:
            self.ui.file_path_window.setPlainText("Выбран файл мощности по пути: " + str(self._filepath_txt))
            self.set_db_graph(self.data_list_db_X, self.data_list_db_Y, False)
        else:
            self.ui.file_path_window.setPlainText("Файл мощности не выбран")
            self.set_db_graph(0, 0, True)

        self.filepath_txt = self._filepath_txt

    def update_sun_activity(self):
        self.ClearSun()

        self.x_values = []

        input_format = '%d-%b-%Y %H:%M:%S'
        output_format = '%Y-%m-%dT%H:%M:%S'

        if not self.filepath_txt[0]:
            self.show_critical_message("Ошибка", "Файл мощности сигнала (.xlsx) не был выбран!")
            return
        else:
            print("Солнечная активность")
            print("Идёт скачивание...")

            # datetime_obj_start = datetime.strptime(self.start_data.strip("'"), input_format)
            start_time = self.to_sunPy_time(self.start_data)

            # datetime_obj_end = datetime.strptime(self.end_data.strip("'"), input_format)
            end_time = self.to_sunPy_time(self.end_data)

            query = Fido.search(a.Time(start_time, end_time), a.Instrument('GOES'))
            files = Fido.fetch(query)

            if files:
                file = files[0]
                goes = ts.TimeSeries(file)
                goes_subset = goes.truncate(start_time, end_time)
                df = goes_subset.to_dataframe()
                self.x_values = df['xrsb'].astype(float).tolist()

                file = files[0]
                goes = ts.TimeSeries(file)
                goes_subset = goes.truncate(start_time, end_time)

                # Преобразуем TimeSeries в DataFrame
                df = goes_subset.to_dataframe()

                # Создаем виджет макета
                layout = pg.GraphicsLayoutWidget(title="Солнечная активность")
                view = layout.addViewBox()
                view.setAspectLocked(True)

                self.set_sun_graph(df.index, df.iloc[:, 0], False)

                print("График успешно построен!")
                print("Start time: " + str(start_time))
                print("End time: " + str(end_time))
                print("Количество строк:", len(self.x_values))
                print("x_values: ", self.x_values)
                time.sleep(1)
                print("_" * 100)
            else:
                print("Нет доступных данных для указанного временного интервала.")

    def compare_click(self):
        self.ClearCompare()

        if not self.filepath_txt[0]:
            self.show_critical_message("Ошибка", "Файл мощности сигнала (.xlsx) не был выбран!")
            self.ui.CorrBox.setPlainText(f"Корреляция: -")
            return

        if not self.x_values:
            self.show_critical_message("Ошибка", "Солнечная активность не была получена!")
            self.ui.CorrBox.setPlainText(f"Корреляция: -")
            return

        self.corr_list_db_Y = []
        self.corr_list_sun_Y = []

        for i in range(min(len(self.data_list_db_X), len(self.x_values))):
            self.corr_list_db_Y.append(self.data_list_db_Y[i])
            self.corr_list_sun_Y.append(self.x_values[i])

        self.corr_list_db_Y = list(map(float, self.corr_list_db_Y))
        self.corr_list_sun_Y = list(map(float, self.corr_list_sun_Y))

        corr = np.corrcoef(self.corr_list_sun_Y, self.corr_list_db_Y)[0, 1]

        if not self.x_values:
            self.show_critical_message("Ошибка", "Солнечная активность не была получена!")
        else:
            if len(self.data_list_db_X) != len(self.x_values):
                self.show_warning_message("Разное количество строк!")
            else:
                self.show_information_message("Успешно", "Данные сопоставлены успешно.")
            self.compare_out(self.corr_list_db_Y, self.corr_list_sun_Y, False)
            print("Корреляция")
            print("Количество объектов:", min(len(self.corr_list_db_Y), len(self.corr_list_sun_Y)))
            print("Длина corr_list_db_Y: ", len(self.corr_list_db_Y))
            print("Длина corr_list_sun_Y: ", len(self.corr_list_sun_Y))
            print("corr_list_db_Y:", self.corr_list_db_Y)
            print("corr_list_sun_Y", self.corr_list_sun_Y)
            print("Коэффициент корреляции:", corr)
            self.ui.CorrBox.setPlainText(f"Корреляция: {corr:.2f}")
            print("_" * 100)

        """
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Ошибка")
            msg.setInformativeText('Не получена солнечная активность!')
            msg.setWindowTitle("Сопоставление данных")
            msg.exec_()
        """

    def show_warning_message(self, text):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Warning)
        msg.setText(text)
        msg.setInformativeText('Часть строк была обрезана, во избежание ошибки!')
        msg.setWindowTitle("Сопоставление данных")
        msg.exec_()

    def show_critical_message(self, text, informative_text):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText(text)
        msg.setInformativeText(informative_text)
        msg.setWindowTitle("Сопоставление данных")
        msg.exec_()

    def show_information_message(self, text, informative_text):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText(text)
        msg.setInformativeText(informative_text)
        msg.setWindowTitle("Сопоставление данных")
        msg.exec_()

    def compare_out(self, activity, power_db, clear):  # Вывод графика зависимости солнечной активности от db
        if not clear:
            poly = np.polyfit(self.corr_list_db_Y, self.corr_list_sun_Y, deg=3)
            line = np.polyval(poly, self.corr_list_db_Y)

            self.ui.graph_out.plot(activity, power_db, pen=None, name="blue", symbol='o')
            self.ui.graph_out.plot(self.corr_list_db_Y, line, pen='g', symbolBrush=1)
        else:
            self.ui.graph_out.clear()

    def set_db_graph(self, time, db, clear):  # Вывод графика зависимости времени от db
        if not clear:
            self.ui.grap_db.plot(time, db)
        else:
            self.ui.grap_db.clear()

    def set_sun_graph(self, activity, time, clear):  # Вывод графика зависимости солнечной активности от времени
        if not clear:
            self.ui.graph_sun.plot(activity, time)
        else:
            self.ui.graph_sun.clear()


app = QtWidgets.QApplication(sys.argv)

window = MainWindow()
window.setFixedSize(997, 486)
window.show()
app.exec()